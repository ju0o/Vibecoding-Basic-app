"""Export CURRICULUM_MASTER.csv + Day1 outcomes to review XLSX (derivative, not SSOT)."""
from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
CSV_PATH = ROOT / "ai-ops" / "curriculum" / "CURRICULUM_MASTER.csv"
OUT = ROOT / "exports" / "curriculum" / "CURRICULUM_MASTER.xlsx"


def style_header(cell) -> None:
    cell.font = Font(name="Arial", bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor="E8EEF7")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )


def style_cell(cell) -> None:
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Curriculum ---
    ws = wb.active
    ws.title = "Curriculum"
    headers = [
        "Course ID",
        "Stage ID",
        "Lesson ID",
        "Order",
        "Lesson Title",
        "Student Question",
        "Why Now",
        "Learning Goal",
        "Learning Outcomes",
        "Practice",
        "Interaction",
        "Assessment",
        "Atlas References",
        "Tool References",
        "Prerequisites",
        "Next Lesson",
        "Source Status",
        "Content Status",
        "Reviewer Status",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        style_header(c)

    row_i = 2
    with CSV_PATH.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            values = [
                row["course_id"],
                row["stage_id"],
                row["lesson_id"],
                int(row["order"]),
                row["lesson_title"],
                row["student_question"].replace("|", "\n"),
                row["why_now"],
                row["learning_goal"],
                row["outcomes"].replace(";", ", "),
                row["practice"],
                row["interaction"],
                row["assessment"],
                row["atlas_refs"],
                row["tool_refs"],
                row["prerequisites"],
                row["next_lesson"],
                row["source_status"],
                row["content_status"],
                row["reviewer_status"],
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row_i, col, v)
                style_cell(c)
            row_i += 1

    last = max(row_i - 1, 1)
    ws.auto_filter.ref = f"A1:S{last}"
    ws.freeze_panes = "A2"
    widths = [22, 18, 18, 8, 36, 28, 28, 28, 28, 40, 40, 40, 14, 22, 14, 12, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    if last >= 2:
        ws.row_dimensions[2].height = 80
    note = ws.cell(
        last + 2,
        1,
        f"SSOT CSV: ai-ops/curriculum/CURRICULUM_MASTER.csv | Generated: {date.today().isoformat()} | DERIVATIVE — edit CSV not this file",
    )
    note.font = Font(name="Arial", size=9, italic=True, color="666666")

    # --- Day 1 Outcomes ---
    ws2 = wb.create_sheet("Day 1 Outcomes")
    h2 = [
        "Outcome ID",
        "Outcome",
        "Required Level",
        "Minimum Completion",
        "Recommended Completion",
        "Evidence",
        "Assessment Method",
        "Deferred Allowed",
        "Notes",
    ]
    for col, h in enumerate(h2, 1):
        style_header(ws2.cell(1, col, h))

    outcomes = [
        (
            "O1",
            "바이브코딩과 전통 코딩 차이를 자기 말로 설명",
            "Assisted",
            "Assisted",
            "Explainable",
            "teach-back 문장",
            "Teach-back / checklist",
            "No",
            "개념",
        ),
        (
            "O2",
            "AI에게 작은 결과물 요청",
            "Independent",
            "Independent",
            "Independent",
            "요청 기록·결과물",
            "수행",
            "No",
            "Path A",
        ),
        (
            "O3",
            "결과 보고 수정 요청 ≥1",
            "Independent",
            "Independent",
            "Independent",
            "수정 전후 화면",
            "수행",
            "No",
            "Path A",
        ),
        (
            "O4",
            "IDE가 작업 공간임을 설명",
            "Assisted",
            "Assisted",
            "Explainable",
            "한 줄 설명",
            "Teach-back",
            "No",
            "",
        ),
        (
            "O5",
            "VS Code 설치 여부 확인",
            "Independent",
            "Independent",
            "Independent",
            "실행 또는 미설치 확인",
            "수행",
            "No",
            "대안 편집기 OK",
        ),
        (
            "O6",
            "Node.js 설치 여부 확인",
            "Independent",
            "Assisted+",
            "Independent",
            "node -v 또는 미설치 문서화",
            "수행",
            "Yes if no admin",
            "Path B",
        ),
        (
            "O7",
            "터미널 열기",
            "Independent",
            "Assisted+",
            "Independent",
            "터미널 화면",
            "수행",
            "No",
            "",
        ),
        (
            "O8",
            "node -v 와 npm -v 실행",
            "Independent",
            "Assisted+",
            "Independent",
            "버전 출력 메모",
            "수행",
            "Yes if no Node",
            "Path B",
        ),
        (
            "O9",
            "제공/로컬 프로젝트 폴더 열기",
            "Assisted",
            "Assisted if env allows",
            "Independent",
            "폴더 연 상태",
            "수행",
            "Yes",
            "examples/day1-first-success",
        ),
        (
            "O10",
            "안내에 따라 개발 서버 실행",
            "Assisted",
            "Assisted if env allows",
            "Independent",
            "npm run dev",
            "수행",
            "Yes",
            "Path B",
        ),
        (
            "O11",
            "브라우저에서 결과 확인",
            "Independent",
            "Independent",
            "Independent",
            "화면 확인",
            "수행",
            "No",
            "Path A 또는 B",
        ),
        (
            "O12",
            "package.json/src/npm install/run dev 기본 설명",
            "Assisted",
            "Assisted",
            "Explainable",
            "한 줄 설명",
            "Teach-back",
            "No",
            "",
        ),
        (
            "O13",
            "오류 메시지 복사 후 AI 전달",
            "Independent",
            "Assisted",
            "Independent",
            "템플릿 사용",
            "시나리오/모의 OK",
            "No",
            "모의 허용",
        ),
    ]
    for r, row in enumerate(outcomes, 2):
        for col, v in enumerate(row, 1):
            style_cell(ws2.cell(r, col, v))
    ws2.auto_filter.ref = f"A1:I{1 + len(outcomes)}"
    ws2.freeze_panes = "A2"
    for i, w in enumerate([10, 42, 14, 18, 20, 18, 16, 14, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    n2 = ws2.cell(
        16,
        1,
        f"Source: DAY1-OUTCOME-CONTRACT.md + assessment | Generated: {date.today().isoformat()}",
    )
    n2.font = Font(name="Arial", size=9, italic=True, color="666666")

    # --- Production Status ---
    ws3 = wb.create_sheet("Production Status")
    for col, h in enumerate(["Item", "Path / Note", "Status", "Next Action"], 1):
        style_header(ws3.cell(1, col, h))
    status_rows = [
        (
            "Student Content",
            "content/courses/vibe-coding-foundation/lessons/01-first-success.md",
            "drafting",
            "Operator read-through",
        ),
        (
            "Instructor Guide",
            "content/instructor/vibe-coding-foundation/01-first-success-instructor.md",
            "drafting",
            "Operator read-through",
        ),
        (
            "Practice",
            "content/practice/vibe-coding-foundation/01-first-success-practice.md",
            "drafting",
            "Align with sample project",
        ),
        (
            "Interaction Spec",
            "content/interactions/vibe-coding-foundation/01-first-success-interaction-spec.md",
            "drafting",
            "Storyboard review",
        ),
        (
            "Assessment",
            "content/assessment/vibe-coding-foundation/01-first-success-assessment.md",
            "drafting",
            "Confirm min complete",
        ),
        (
            "Source Pack",
            "ai-ops/reports/research/DAY1-SOURCE-PACK.md",
            "partial",
            "Re-check LTS on publish",
        ),
        (
            "Independent Review",
            "ai-ops/reports/DAY1-INDEPENDENT-REVIEW.md",
            "pass_with_notes",
            "Human re-stamp optional",
        ),
        (
            "Curriculum CSV SSOT",
            "ai-ops/curriculum/CURRICULUM_MASTER.csv",
            "drafting",
            "Add Day 2 after approve",
        ),
        (
            "Curriculum XLSX export",
            "exports/curriculum/CURRICULUM_MASTER.xlsx",
            "generated",
            "Not SSOT",
        ),
        (
            "Student DOCX export",
            "exports/student/DAY1-*.docx",
            "generated",
            "Not SSOT",
        ),
        (
            "Instructor DOCX export",
            "exports/instructor/DAY1-*.docx",
            "generated",
            "Not SSOT",
        ),
        (
            "Sample Project",
            "examples/day1-first-success/",
            "verified_local",
            "Keep zero external deps",
        ),
        (
            "Website Connection",
            "src/app / student routes",
            "not_started",
            "After operator APPROVE",
        ),
        (
            "Animation Implementation",
            "UI from interaction spec",
            "not_started",
            "After content approve",
        ),
        (
            "QA",
            "ai-ops/reports/DAY1-OPERATOR-PACKAGE-QA.md",
            "pending",
            "Operator package QA",
        ),
    ]
    for r, row in enumerate(status_rows, 2):
        for col, v in enumerate(row, 1):
            style_cell(ws3.cell(r, col, v))
    ws3.auto_filter.ref = f"A1:D{1 + len(status_rows)}"
    ws3.freeze_panes = "A2"
    for i, w in enumerate([28, 55, 16, 28], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    n3 = ws3.cell(
        18,
        1,
        f"Generated: {date.today().isoformat()} | No fake completion % formulas",
    )
    n3.font = Font(name="Arial", size=9, italic=True, color="666666")

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
